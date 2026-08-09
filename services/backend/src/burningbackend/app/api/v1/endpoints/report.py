from burningbackend.app.models.history import History
from burningbackend.app.models.movie import Movie
import burningbackend.app.api.v1.endpoints.history as history

from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill, numbers
from openpyxl.utils import get_column_letter

from io import BytesIO

from fastapi import APIRouter
from fastapi import HTTPException
from starlette.responses import StreamingResponse

router = APIRouter()


def _style_header_row(ws, row, col_start, col_end, fill_color="2E3440"):
    header_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    header_font = Font(name='Calibri', bold=True, color="ECEFF4", size=11)
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _style_data_cell(ws, row, col, value, is_currency=False, is_bold=False, align="left"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name='Calibri', bold=is_bold, size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if is_currency and isinstance(value, (int, float)):
        cell.number_format = '#,##0.00 €'
    return cell


def _add_thin_border(ws, row, col_start, col_end):
    thin = Side(style='thin', color="D8DEE9")
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = Border(bottom=thin)


def _auto_column_width(ws, min_width=10, max_width=30):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        adjusted = min(max(max_len + 3, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


@router.get("/report", response_description="Excel report retrieved")
async def get_report(movie: str):
    selected_movie = await Movie.find_one({"name": movie})
    if not selected_movie:
        raise HTTPException(status_code=404, detail=f"Movie '{movie}' not found")

    data_sold = await History.find({"movie": movie, "cancellation": False, "isteam": False}).to_list()
    team_data = await History.find({"movie": movie, "cancellation": False, "isteam": True}).to_list()
    cancelled_data = await History.find({"movie": movie, "cancellation": True}).to_list()

    total_sold = await history.get_total(movie, False, False, True)
    total_sold_team = await history.get_total(movie, True, False, True)
    total_without_pfand = await history.get_total(movie, False, False, False)

    def summarize_products(orders):
        product_summary = {}
        for order in orders:
            for product in order.products:
                name = product.name
                amount = product.amount
                price = product.price
                total_price = price * amount
                if name in product_summary:
                    product_summary[name]['amount'] += amount
                    product_summary[name]['total'] += total_price
                else:
                    product_summary[name] = {
                        'amount': amount,
                        'total': total_price,
                        'price': price
                    }
        return product_summary

    def safe_get(d, key, field):
        return d[key][field] if key in d else 0

    products_dict = summarize_products(data_sold)
    team_products_dict = summarize_products(team_data)
    cancelled_products_dict = summarize_products(cancelled_data)

    all_product_names = sorted(set(
        list(products_dict.keys()) + list(team_products_dict.keys())
    ))

    wb = Workbook()

    # ── Sheet 1: Summary ──
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_properties.tabColor = "FF6B6B"

    title_font = Font(name='Calibri', bold=True, size=14, color="FF6B6B")
    section_font = Font(name='Calibri', bold=True, size=12, color="4ECDC4")
    label_font = Font(name='Calibri', size=10)
    value_font = Font(name='Calibri', bold=True, size=10)
    currency_fmt = '#,##0.00 €'

    # Movie header
    ws.cell(row=1, column=1, value="Burning Register — Report").font = title_font
    ws.merge_cells('A1:D1')

    row = 3
    ws.cell(row=row, column=1, value="Movie Information").font = section_font
    row += 1
    for label, val in [
        ("Movie", selected_movie.name),
        ("Room", selected_movie.room),
        ("Date & Time", str(selected_movie.datetime)),
        ("Total Orders", len(data_sold) + len(team_data)),
        ("Cancelled Orders", len(cancelled_data)),
    ]:
        ws.cell(row=row, column=1, value=label).font = label_font
        ws.cell(row=row, column=2, value=val).font = value_font
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Financial Summary").font = section_font
    row += 1

    financial_data = [
        ("Total Revenue (incl. Pfand)", total_sold),
        ("Revenue without Pfand", total_without_pfand),
        ("Pfand Collected", safe_get(products_dict, "Pfand", "total")),
        ("Products Sold (excl. Tickets & Pfand)", total_sold - safe_get(products_dict, "Ticket", "total") - safe_get(products_dict, "Clubkarte", "total") - safe_get(products_dict, "Pfand", "total")),
        ("Ticket Revenue", safe_get(products_dict, "Ticket", "total")),
        ("Clubkarten Revenue", safe_get(products_dict, "Clubkarte", "total")),
        ("Team Consumption Total", total_sold_team),
    ]

    for label, val in financial_data:
        ws.cell(row=row, column=1, value=label).font = label_font
        c = ws.cell(row=row, column=2, value=val)
        c.font = value_font
        c.number_format = currency_fmt
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Ticket Summary").font = section_font
    row += 1

    ticket_data = [
        ("Paid Tickets", safe_get(products_dict, "Ticket", "amount")),
        ("Free Tickets (Freitickets)", safe_get(products_dict, "Freiticket", "amount")),
        ("Total Visitors", safe_get(products_dict, "Ticket", "amount") + safe_get(products_dict, "Freiticket", "amount")),
    ]
    for label, val in ticket_data:
        ws.cell(row=row, column=1, value=label).font = label_font
        ws.cell(row=row, column=2, value=val).font = value_font
        row += 1

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 18

    # ── Sheet 2: Products Breakdown ──
    ws2 = wb.create_sheet("Products")
    ws2.sheet_properties.tabColor = "4ECDC4"

    headers = ["Product", "Qty Sold", "Unit Price", "Total Revenue", "Team Qty", "Team Price", "Team Total"]
    for i, h in enumerate(headers, 1):
        ws2.cell(row=1, column=i, value=h)
    _style_header_row(ws2, 1, 1, len(headers))

    data_row = 2
    grand_total = 0
    grand_team_total = 0
    for name in all_product_names:
        amount = safe_get(products_dict, name, 'amount')
        price = safe_get(products_dict, name, 'price')
        total = safe_get(products_dict, name, 'total')
        team_amount = safe_get(team_products_dict, name, 'amount')
        team_price = safe_get(team_products_dict, name, 'price')
        team_total = team_price * team_amount

        _style_data_cell(ws2, data_row, 1, name, align="left", is_bold=True)
        _style_data_cell(ws2, data_row, 2, amount, align="center")
        _style_data_cell(ws2, data_row, 3, price, is_currency=True, align="right")
        _style_data_cell(ws2, data_row, 4, total, is_currency=True, align="right", is_bold=True)
        _style_data_cell(ws2, data_row, 5, team_amount, align="center")
        _style_data_cell(ws2, data_row, 6, team_price, is_currency=True, align="right")
        _style_data_cell(ws2, data_row, 7, team_total, is_currency=True, align="right")

        grand_total += total
        grand_team_total += team_total
        data_row += 1

    # Totals row
    _add_thin_border(ws2, data_row - 1, 1, 7)
    _style_data_cell(ws2, data_row, 1, "TOTAL", is_bold=True)
    _style_data_cell(ws2, data_row, 4, grand_total, is_currency=True, is_bold=True, align="right")
    _style_data_cell(ws2, data_row, 7, grand_team_total, is_currency=True, is_bold=True, align="right")

    total_fill = PatternFill(start_color="3B4252", end_color="3B4252", fill_type="solid")
    for col in range(1, 8):
        ws2.cell(row=data_row, column=col).fill = total_fill

    _auto_column_width(ws2)

    # ── Sheet 3: Order Log ──
    ws3 = wb.create_sheet("Orders")
    ws3.sheet_properties.tabColor = "64B5F6"

    order_headers = ["#", "Timestamp", "Products", "Team", "Cancelled", "Total"]
    for i, h in enumerate(order_headers, 1):
        ws3.cell(row=1, column=i, value=h)
    _style_header_row(ws3, 1, 1, len(order_headers))

    all_orders = data_sold + team_data + cancelled_data
    all_orders.sort(key=lambda o: o.timestamp)

    for idx, order in enumerate(all_orders, 1):
        r = idx + 1
        products_str = ", ".join(f"{p.name} x{p.amount}" for p in order.products)
        _style_data_cell(ws3, r, 1, idx, align="center")
        _style_data_cell(ws3, r, 2, str(order.timestamp), align="left")
        _style_data_cell(ws3, r, 3, products_str, align="left")
        _style_data_cell(ws3, r, 4, "Yes" if order.isteam else "No", align="center")

        cancel_cell = _style_data_cell(ws3, r, 5, "Yes" if order.cancellation else "No", align="center")
        if order.cancellation:
            cancel_cell.font = Font(name='Calibri', color="FF5252", bold=True, size=10)

        _style_data_cell(ws3, r, 6, order.total, is_currency=True, align="right")

    _auto_column_width(ws3, min_width=8)
    ws3.column_dimensions['C'].width = 40

    # Save
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response.headers["Content-Disposition"] = f"attachment; filename={selected_movie.name}.xlsx"

    return response
